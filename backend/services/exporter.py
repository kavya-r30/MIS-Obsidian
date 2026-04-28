import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def generate_excel(transactions: list[dict], summary: dict) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    ws_summary.append(["Metric", "Value"])
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
    for k, v in summary.items():
        ws_summary.append([k, v])

    if transactions:
        ws_tx = wb.create_sheet("Transactions")
        ws_tx.append(list(transactions[0].keys()))
        for cell in ws_tx[1]:
            cell.font = header_font
            cell.fill = header_fill
        for tx in transactions:
            ws_tx.append(list(tx.values()))

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def generate_pdf_html(transactions: list[dict], summary: dict, narrative: str) -> str:
    rows = "".join(
        f"<tr>{''.join(f'<td>{v if v is not None else chr(8212)}</td>' for v in tx.values())}</tr>"
        for tx in transactions
    )
    headers = "".join(f"<th>{k}</th>" for k in (transactions[0].keys() if transactions else []))
    kpi_cards = "".join(
        f'<div class="kpi"><div class="kpi-label">{k}</div><div class="kpi-value">{v}</div></div>'
        for k, v in summary.items()
    )
    return f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  body {{ font-family: 'Helvetica Neue', Arial, sans-serif; padding: 30px; color: #2c3e50; }}
  h1 {{ font-size: 22px; margin-bottom: 4px; }}
  .subtitle {{ color: #7f8c8d; font-size: 13px; margin-bottom: 24px; }}
  .kpi-row {{ display: flex; flex-wrap: wrap; gap: 12px; margin: 20px 0; }}
  .kpi {{ background: #3498db; color: white; padding: 14px 20px; border-radius: 8px; min-width: 120px; }}
  .kpi-label {{ font-size: 11px; opacity: 0.85; text-transform: uppercase; }}
  .kpi-value {{ font-size: 24px; font-weight: bold; margin-top: 4px; }}
  .narrative {{ background: #f0f4f8; padding: 16px 20px; border-radius: 8px; margin: 20px 0; font-size: 13px; line-height: 1.6; border-left: 4px solid #3498db; }}
  .narrative h2 {{ font-size: 14px; margin: 0 0 8px 0; color: #2980b9; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 20px; font-size: 12px; }}
  th {{ background: #2c3e50; color: white; padding: 9px 10px; text-align: left; }}
  td {{ border-bottom: 1px solid #ecf0f1; padding: 8px 10px; }}
  tr:nth-child(even) td {{ background: #f8f9fa; }}
</style>
</head>
<body>
  <h1>MIS Transaction Validation Report</h1>
  <div class="subtitle">Generated on {__import__('datetime').datetime.now().strftime('%d %B %Y, %H:%M')}</div>
  <div class="kpi-row">{kpi_cards}</div>
  <div class="narrative"><h2>AI Executive Summary</h2><p>{narrative}</p></div>
  <h2 style="font-size:15px;margin-top:28px;">Transaction Details</h2>
  <table><thead><tr>{headers}</tr></thead><tbody>{rows}</tbody></table>
</body>
</html>"""
